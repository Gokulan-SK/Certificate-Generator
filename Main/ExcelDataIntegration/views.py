from django.shortcuts import render
from openpyxl import load_workbook
from .models import IncompleteGeneration
from django.contrib.auth.decorators import login_required



def import_form(request):
    if request.method=='POST':
        excel_file=request.FILES['excel_file']
        wb=load_workbook(excel_file)
        ws=wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, course, duration, email = row
            IncompleteGeneration.objects.create(name=name, course=course, duration=duration, email=email)

        return render(request, 'dashboard.html')

    return render(request, 'import_form.html')
